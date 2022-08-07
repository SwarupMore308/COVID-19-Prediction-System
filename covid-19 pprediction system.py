from tkinter import *
from tkinter import ttk
import tkinter.messagebox
import numpy as np
import pandas as pd
import time
import openpyxl
from win32com import client
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


res1 = []
# List of the symptoms is listed here in list l1.


l1 = ['fever','dry_cough','tiredness','aches_and_pains','diarrhoea','sore_throat','conjunctivitis','headache','loss_of_taste_or_smell','rash_on_skin','discolouration_of_fingers_or_toes','difficulty_breathing_or_shortness_of_breath','chest_pain_or_pressure','loss_of_speech_or_movement']

# list of disease

ldisease = ['corona_virus', 'simple_cough', 'viral_infection', 'throat_infection', 'lac_of_vitamins' ]

l2 = []
for i in range(0, len(l1)):
    l2.append(0)

# importing csv file
df = pd.read_csv("trdt.csv")

# Replace the values in the imported file by pandas by the inbuilt function replace in pandas.
df.replace({'prognosis': {'corona_virus': 0, 'simple_cough': 1, 'viral_infection': 2, 'throat_infection': 3,
                          'lac_of_vitamins': 4}}, inplace=True)

# check the df

# print(df.head())
X = df[l1]

# print(X)
y = df[["prognosis"]]
np.ravel(y)

# print(y)

# Read a csv named Testing.csv
tr = pd.read_csv("trdt2.csv")

# Use replace method in pandas.
tr.replace({'prognosis': {'corona_virus': 0, 'simple_cough': 1, 'viral_infection': 2, 'throat_infection': 3,
                          'lac_of_vitamins': 4}}, inplace=True)
X_test = tr[l1]
y_test = tr[["prognosis"]]

# print(y_test)
np.ravel(y_test)

# creating GUI

root = Tk()
root.title("Covid-19 Prediction system")
root.geometry("1640x850")
root.configure(bg="navy")

sugaro = ['Sugar', 'No_Sugar']

title = StringVar("")
f_name = StringVar("")
m_name = StringVar("")
l_name = StringVar("")
age = StringVar("")
gender = StringVar("")
email = StringVar("")
birthdate = StringVar("")
contact = StringVar("")
address = StringVar("")
pin_code = StringVar("")

res = ""

Date = StringVar()
Date.set(time.strftime("%d/%m/%y"))


def next_page():
    root.destroy()
    root1 = Tk()
    root1.title("Covid-19 Prediction System")
    root1.geometry("1600x850")
    root1.configure(bg="navy")

    # algorithms:

    def DecisionTree():
        from sklearn import tree
        clf3 = tree.DecisionTreeClassifier()
        clf3 = clf3.fit(X, y)
        from sklearn.metrics import accuracy_score
        y_pred = clf3.predict(X_test)
        print(accuracy_score(y_test, y_pred))
        print(accuracy_score(y_test, y_pred, normalize=False))
        psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get(), Symptom6.get(),
                     Symptom7.get(),
                     Symptom8.get(), Symptom9.get(), Symptom10.get(), Symptom11.get(), Symptom12.get(), Symptom13.get(),
                     Symptom14.get()]
        for k in range(0, len(l1)):
            for z in psymptoms:
                if z == l1[k]:
                    l2[k] = 1
        inputtest = [l2]
        predict = clf3.predict(inputtest)
        predicted = predict[0]
        h = 'no'
        for a in range(0, len(ldisease)):
            if predicted == a:
                h = 'yes'
                break
        if h == 'yes':
            t1.delete("1.0", END)
            t1.insert(END, ldisease[a])
            res1.append(t1.get("1.0", END))
            print(res1)

        else:
            t1.delete("1.0", END)
            t1.insert(END, "Not Found")



    def randomforest():
        global notify
        from sklearn.ensemble import RandomForestClassifier
        clf4 = RandomForestClassifier()
        clf4 = clf4.fit(X, np.ravel(y))

        # calculating accuracy

        from sklearn.metrics import accuracy_score
        y_pred = clf4.predict(X_test)
        print(accuracy_score(y_test, y_pred))
        print(accuracy_score(y_test, y_pred, normalize=False))

        psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get(), Symptom6.get(),
                     Symptom7.get(),
                     Symptom8.get(), Symptom9.get(), Symptom10.get(), Symptom11.get(), Symptom12.get(), Symptom13.get(),
                     Symptom14.get()]

        for k in range(0, len(l1)):
            for z in psymptoms:
                if z == l1[k]:
                    l2[k] = 1

        inputtest = [l2]
        predict = clf4.predict(inputtest)
        predicted = predict[0]
        h = 'no'
        for a in range(0, len(ldisease)):
            if predicted == a:
                h = 'yes'
                break
        if h == 'yes':
            t2.delete("1.0", END)
            t2.insert(END, ldisease[a])
        else:
            t2.delete("1.0", END)
            t2.insert(END, "Not Found")

        res2 = t2.get("1.0",END)

    # exit function for next page :
    def exit1():
        e = tkinter.messagebox.askyesno("Covid-19 Prediction System", "Confirm you want to Exit")
        if e > 0:
            root1.destroy()
            return

    def reportgenrate():
        wb = openpyxl.load_workbook(
            "D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report_Sample_1.xlsx")
        sh1 = wb['sheet1']
        row = sh1.max_row
        column = sh1.max_column

        name_of_p = title.get() + " " + f_name.get() + " " + m_name.get() + " " + l_name.get()
        gender_p = gender.get()
        address_p = address.get()
        result = str(res1[0])
        result1 = " "

        print(result)
        sh1.cell(row=3, column=3, value=name_of_p)
        sh1.cell(row=4, column=3, value=gender_p)
        sh1.cell(row=5, column=3, value=birthdate.get())
        sh1.cell(row=6, column=3, value=age.get())
        sh1.cell(row=7, column=3, value=Date.get())
        sh1.cell(row=3, column=8, value=address_p)
        sh1.cell(row=4, column=8, value=contact.get())
        sh1.cell(row=5, column=8, value=email.get())

        if result == 'corona_virus':
            sh1.cell(row=9, column=7, value="Positive")
            print(1)
        else:
            sh1.cell(row=9, column=7, value="Negative")
            print(2)
        savename = name_of_p.upper()
        wb.save("D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report in Excel\\"+str(savename)+".xlsx")

        xlApp = client.Dispatch("Excel.Application")
        books = xlApp.Workbooks.Open(
            "D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report in Excel\\"+str(savename)+".xlsx")
        ws = books.Worksheets[0]
        ws.Visible = 1
        ws.ExportAsFixedFormat(0, 'D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report in PDF\\'+str(savename)+'.xlsx')

    def send_mail():
        fromaddr = "dvmssoftwares.official@gmail.com"
        toaddr = str(email.get())

        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        msg['Subject'] = "COVID-19 Prediction Result"
        body = "We predicted "+str(res1[0])+" disease with the data you have provided us. their is a report attached with this mail. Have a look "
        msg.attach(MIMEText(body, 'plain'))
        name_of_p = title.get() + " " + f_name.get() + " " + m_name.get() + " " + l_name.get()
        filename = name_of_p+".pdf"
        path="D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report in PDF\\"+str(name_of_p.upper())+".xlsx.pdf"
        attachment = open(path, "rb")
        p = MIMEBase('application', 'octet-stream')
        p.set_payload((attachment).read())
        encoders.encode_base64(p)

        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(p)

        # creates SMTP session
        s = smtplib.SMTP('smtp.gmail.com', 587)

        # start TLS for security
        s.starttls()

        # Authentication
        s.login(fromaddr, "Dvms535326")

        # Converts the Multipart msg into a string
        text = msg.as_string()

        # sending the mail
        s.sendmail(fromaddr, toaddr, text)

        # terminating the session
        s.quit()


    def last_page():
        root1.destroy()
        root2 = Tk()
        root2.title("Covid-19 Prediction system")
        root2.geometry("1600x850")
        root2.configure(bg="White")


        def exit3():
            e1 = tkinter.messagebox.askyesno("Covid-19 Prediction System", "Confirm you want to Exit")
            if e1 > 0:
                root2.destroy()
                return

        t3 = Text(height=30, width=80, font="Times 16 bold")
        t3.grid(row=1, column=0, padx=25, pady=15)

        t3.insert(END, "\n\t\t\t\t" + "REPORT OF PATIENT :")
        t3.insert(END,
                  "\n\n\n\t" + "Name Of Patient  :           " + title.get() + "\t" + f_name.get() + "\t" + m_name.get() + "\t" + l_name.get())
        t3.insert(END, "\n\n\n\t" + "Age Of Patient :             " + age.get())
        t3.insert(END, "\n\n\n\t" + "Gender Of Patient  :         " + gender.get())
        t3.insert(END, "\n\n\n\t" + "Birth date Of Patient  :     " + birthdate.get())
        t3.insert(END, "\n\n\n\t" + "Address Of Patient  :        " + address.get())
        t3.insert(END, "\n\n\n\t" + "Contact No. Of Patient  :    " + contact.get())
        t3.insert(END, "\n\n\n\t" + "Email-Id Of Patient  :       " + email.get())
        t3.insert(END, "\n\n\n\t" + "Date   :       " + Date.get())
        t3.insert(END, "\n\n\n\t" + "Predicted Result Of Patient : P1. "+str(res1[0]))

        safety = Label(text="Safety Measures And Precautions :", font="Times 17 bold", background="White")
        safety.grid(row=0, column=1, sticky=W, padx=100)

        t4 = Text(height=27, width=50, font="Times 16 ")
        t4.grid(row=1, column=1, padx=15, pady=15)

        t4.insert(END,
                  "\n1] Clean your hands often. Use soap and water, or an alcohol-based hand rub.\n\n2] Maintain a safe distance from anyone who is coughing or sneezing.\n")
        t4.insert(END, "\n3]Wear a mask when physical distancing is not possible.\n")
        t4.insert(END, "\n4] Do not touch your eyes, nose, or mouth.\n\n5] Stay home if you feel unwell.\n")
        t4.insert(END,
                  "\n6] Keep up to date on the latest information from trusted sources, such as WHO or your local and national health authorities.")


        exit_btn2 = Button(root2, text=" EXIT ", font="Times 16 bold", command=exit3, width=20, foreground="White",
                           background="navy")
        exit_btn2.grid(row=2, column=1, padx=20)

        reportgenrate()
        send_mail()
        root2.mainloop()



    tf = Frame(width=1200, relief=RIDGE)
    tf.configure(bg="White")
    tf.grid()

    select_lbl = Label(tf, text="Please Select Symptoms :", font="Times 17 bold", bg="White")
    select_lbl.grid(row=0, column=0, sticky=W)

    Symptom1 = StringVar()
    Symptom1.set("Select Here")

    Symptom2 = StringVar()
    Symptom2.set("Select Here")

    Symptom3 = StringVar()
    Symptom3.set("Select Here")

    Symptom4 = StringVar()
    Symptom4.set("Select Here")

    Symptom5 = StringVar()
    Symptom5.set("Select Here")

    Symptom6 = StringVar()
    Symptom6.set("Select Here")

    Symptom7 = StringVar()
    Symptom7.set("Select Here")

    Symptom8 = StringVar()
    Symptom8.set("Select Here")

    Symptom9 = StringVar()
    Symptom9.set("Select Here")

    Symptom10 = StringVar()
    Symptom10.set("Select Here")

    Symptom11 = StringVar()
    Symptom11.set("Select Here")

    Symptom12 = StringVar()
    Symptom12.set("Select Here")

    Symptom13 = StringVar()
    Symptom13.set("Select Here")

    Symptom14 = StringVar()
    Symptom14.set("Select Here")

    # creating option boxes to select symptoms

    OPTIONS = sorted(l1)

    S1 = OptionMenu(tf, Symptom1, *OPTIONS)
    S1.config(width=30, font="Arial 10 bold")
    S1.grid(row=2, column=1, padx=10)

    S2 = OptionMenu(tf, Symptom2, *OPTIONS)
    S2.config(width=30, font="Arial 10 bold")
    S2.grid(row=3, column=1, padx=10)

    S3 = OptionMenu(tf, Symptom3, *OPTIONS)
    S3.config(width=30, font="Arial 10 bold")
    S3.grid(row=4, column=1, padx=10)

    S4 = OptionMenu(tf, Symptom4, *OPTIONS)
    S4.config(width=30, font="Arial 10 bold")
    S4.grid(row=5, column=1, padx=10)

    S5 = OptionMenu(tf, Symptom5, *OPTIONS)
    S5.config(width=30, font="Arial 10 bold")
    S5.grid(row=6, column=1, padx=10)

    S6 = OptionMenu(tf, Symptom6, *OPTIONS)
    S6.config(width=30, font="Arial 10 bold")
    S6.grid(row=7, column=1, padx=10)

    S7 = OptionMenu(tf, Symptom7, *OPTIONS)
    S7.config(width=30, font="Arial 10 bold")
    S7.grid(row=8, column=1, padx=10)

    S8 = OptionMenu(tf, Symptom8, *OPTIONS)
    S8.config(width=30, font="Arial 10 bold")
    S8.grid(row=9, column=1, padx=10)

    S9 = OptionMenu(tf, Symptom9, *OPTIONS)
    S9.config(width=30, font="Arial 10 bold")
    S9.grid(row=10, column=1, padx=10)

    S10 = OptionMenu(tf, Symptom10, *OPTIONS)
    S10.config(width=30, font="Arial 10 bold")
    S10.grid(row=2, column=3, padx=10)

    S11 = OptionMenu(tf, Symptom11, *OPTIONS)
    S11.config(width=30, font="Arial 10 bold")
    S11.grid(row=3, column=3, padx=10)

    S12 = OptionMenu(tf, Symptom12, *OPTIONS)
    S12.config(width=30, font="Arial 10 bold")
    S12.grid(row=4, column=3, padx=10)

    S13 = OptionMenu(tf, Symptom13, *OPTIONS)
    S13.config(width=30, font="Arial 10 bold")
    S13.grid(row=5, column=3, padx=10)

    S14 = OptionMenu(tf, Symptom14, *OPTIONS)
    S14.config(width=30, font="Arial 10 bold")
    S14.grid(row=6, column=3, padx=10)

    # creating symptom labels :

    ls1 = Label(tf, text="Symptom 1 :", font="Times 15 bold", background="White")
    ls1.grid(row=2, column=0, sticky=W, padx=20, pady=30)

    ls2 = Label(tf, text="Symptom 2 :", font="Times 15 bold", background="White")
    ls2.grid(row=3, column=0, sticky=W, padx=20, pady=27)

    ls3 = Label(tf, text="Symptom 3 :", font="Times 15 bold", background="White")
    ls3.grid(row=4, column=0, sticky=W, padx=20, pady=27)

    ls4 = Label(tf, text="Symptom 4 :", font="Times 15 bold", background="White")
    ls4.grid(row=5, column=0, sticky=W, padx=20, pady=27)

    ls5 = Label(tf, text="Symptom 5 :", font="Times 15 bold", background="White")
    ls5.grid(row=6, column=0, sticky=W, padx=20, pady=27)

    ls6 = Label(tf, text="Symptom 6 :", font="Times 15 bold", background="White")
    ls6.grid(row=7, column=0, sticky=W, padx=20, pady=27)

    ls7 = Label(tf, text="Symptom 7 :", font="Times 15 bold", background="White")
    ls7.grid(row=8, column=0, sticky=W, padx=20, pady=27)

    ls8 = Label(tf, text="Symptom 8 :", font="Times 15 bold", background="White")
    ls8.grid(row=9, column=0, sticky=W, padx=20, pady=27)

    ls9 = Label(tf, text="Symptom 9 :", font="Times 15 bold", background="White")
    ls9.grid(row=10, column=0, sticky=W, padx=20, pady=27)

    ls10 = Label(tf, text="Symptom 10 :", font="Times 15 bold", background="White")
    ls10.grid(row=2, column=2, sticky=W, padx=30, pady=27)

    ls11 = Label(tf, text="Symptom 11 :", font="Times 15 bold", background="White")
    ls11.grid(row=3, column=2, sticky=W, padx=30, pady=27)

    ls12 = Label(tf, text="Symptom 12 :", font="Times 15 bold", background="White")
    ls12.grid(row=4, column=2, sticky=W, padx=30, pady=27)

    ls13 = Label(tf, text="Symptom 13 :", font="Times 15 bold", background="White")
    ls13.grid(row=5, column=2, sticky=W, padx=30, pady=27)

    ls14 = Label(tf, text="Symptom 14 :", font="Times 15 bold", background="White")
    ls14.grid(row=6, column=2, sticky=W, padx=30, pady=27)

    bp = Label(tf, text="Blood Pressure :", font="Times 15 bold", background="White")
    bp.grid(row=7, column=2, sticky=W, padx=20, pady=27)
    cbobp = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=28)
    cbobp['value'] = ('', 'blood_pressure', 'No_blood_pressure')
    cbobp.current(0)
    cbobp.grid(row=7, column=3, sticky=W, padx=10)

    suger = Label(tf, text="sugar :", font="Times 15 bold", background="White")
    suger.grid(row=8, column=2, sticky=W, padx=30, pady=27)
    cbosuger = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=28)
    cbosuger['value'] = ('', 'Diabetes', 'No Diabetes')
    cbosuger.current(0)
    cbosuger.grid(row=8, column=3, sticky=W, padx=10)

    heartdisease = Label(tf, text="Any heart disease:", font="Times 15 bold", background="White")
    heartdisease.grid(row=2, column=4, sticky=W, padx=30, pady=27)
    heartdiseaseo = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=25)
    heartdiseaseo['value'] = ('', 'Yes', 'No')
    heartdiseaseo.current(0)
    heartdiseaseo.grid(row=2, column=5, sticky=W, padx=10)

    anyother = Label(tf, text="Any other disease:", font="Times 15 bold", background="White")
    anyother.grid(row=3, column=4, sticky=W, padx=28, pady=27)
    anyotherd = Entry(tf, font="Arial 12 bold", width=25)
    anyotherd.grid(row=3, column=5, pady=10)

    days = Label(tf, text="Days you're suffering :", font="Times 15 bold", background="White")
    days.grid(row=4, column=4, sticky=W, padx=30, pady=27)
    days = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=25)
    days['value'] = (
        '', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19',
        '20')
    days.current(0)
    days.grid(row=4, column=5, sticky=W, padx=10)

    contactwith = Label(tf, text="Contact with any patient:", font="Times 14 bold", background="White")
    contactwith.grid(row=5, column=4, sticky=W, padx=30, pady=27)
    contactwithc = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=25)
    contactwithc['value'] = ('', 'Yes', 'No')
    contactwithc.current(0)
    contactwithc.grid(row=5, column=5, sticky=W, padx=10)

    reference = Label(tf, text="Reference of :", font="Times 15 bold", background="White")
    reference.grid(row=6, column=4, sticky=W, padx=30, pady=27)
    referencec = ttk.Combobox(tf, font="Times 12 bold", state='readonly', width=25)
    referencec['value'] = ('', 'Doctor', 'Medical', 'Personally')
    referencec.current(0)
    referencec.grid(row=6, column=5, sticky=W, padx=10)

    resdecision_tree = Label(tf, text="DecisionTree :", font="Times 17 bold", background="White")
    resdecision_tree.grid(row=9, column=2, sticky=W, padx=10, pady=22)

    resrandom_forest = Label(tf, text="RandomForest :", font="Times 17 bold", background="White")
    resrandom_forest.grid(row=10, column=2, sticky=W, padx=10, pady=22)

    # creating text field :

    t1 = Text(tf, height=1, width=25, font="Times 14 bold")
    t1.grid(row=9, column=3, padx=10, pady=17)

    t2 = Text(tf, height=1, width=25, font="Times 14 bold")
    t2.grid(row=10, column=3, padx=10)

    # creating buttons :

    pred1btn = Button(tf, text="Prediction 1", font="Times 16 bold", background="navy", command=DecisionTree,
                      foreground="White", width=20)
    pred1btn.grid(row=9, column=4, padx=5, pady=22)

    pred2btn = Button(tf, text="Prediction 2", font="Times 16 bold", background="navy", command=randomforest,
                      foreground="White", width=20)
    pred2btn.grid(row=10, column=4, padx=10, pady=22)

    exit_button = Button(tf, text=" EXIT ", font="Times 15 bold", command=exit1, background="navy", foreground="White",
                         width=20)
    exit_button.grid(row=10, column=5, padx=5, pady=5)

    next_btn = Button(tf, text=" NEXT ", font="Times 15 bold", background="navy", foreground="White", command=last_page,
                      width=20)
    next_btn.grid(row=9, column=5, padx=5, pady=5)

    root1.mainloop()


def exit2():
    e3 = tkinter.messagebox.askyesno("Covid-19 Prediction System", "Confirm you want to Exit")
    if e3 > 0:
        root.destroy()
        return


tf1 = Frame(root, width=1180, bd=20, padx=20)
tf1.configure(bg="navy")
tf1.grid()

w_label = Label(tf1, text="Welcome to Covid-19 Prediction System", font="Times  32 bold", foreground="White",
                background="navy")
w_label.grid(padx=200, pady=10)

tf2 = Frame(root, width=1180, bd=20, padx=20)
tf2.configure(bg="White")
tf2.grid()

detail_lbl = Label(tf2, text="Enter Your Details :", font="Times  17 bold", background="White")
detail_lbl.grid(row=0, column=0, sticky=W, padx=0, pady=20)

name_lbl = Label(tf2, text="Personal Details: ", font="Times  16 bold", background="White")
name_lbl.grid(row=1, column=0, sticky=W, padx=0, pady=25)

l_title = Label(tf2, font="Times 13 bold", text="Title :", padx=5, pady=2, background="White")
l_title.grid(row=2, column=0, sticky=W)
cbomt = ttk.Combobox(tf2, font="Times 12 bold", state='readonly', textvariable=title, width=23)
cbomt['value'] = ('', 'Mr.', 'Mrs.')
cbomt.current(0)
cbomt.grid(row=3, column=0, sticky=W)

n1 = Label(tf2, text="Enter Your First Name :", font="Times 13 bold", background="White")
n1.grid(row=2, column=1, sticky=W, padx=10)
# entry box for name:
ne1 = Entry(tf2, font="Arial 12 bold", textvariable=f_name, width=23)
ne1.grid(row=3, column=1, sticky=W, padx=20)

n2 = Label(tf2, text="Enter Your Middle Name :", font="Times 13 bold", background="White")
n2.grid(row=2, column=2, sticky=W, pady=10)
ne2 = Entry(tf2, font="Arial 12 bold", textvariable=m_name, width=23)
ne2.grid(row=3, column=2, sticky=W)

n3 = Label(tf2, text="Enter Your Last Name :", font="Times 13 bold", background="White")
n3.grid(row=4, column=0, sticky=W, pady=10)
ne3 = Entry(tf2, font="Arial 12 bold", textvariable=l_name, width=23)
ne3.grid(row=5, column=0, pady=10)

l_gender = Label(tf2, font="Times 13 bold", text="Gender :", padx=5, pady=2, background="White")
l_gender.grid(row=4, column=1, sticky=W, padx=20, pady=10)
cbomt1 = ttk.Combobox(tf2, font="Arial 12 bold", state='readonly', textvariable=gender, width=23)
cbomt1['value'] = ('', 'Male.', 'Female.', 'Other.')
cbomt1.current(0)
cbomt1.grid(row=5, column=1, pady=20, padx=20)

splang = Label(tf2, text="Primary Spoken Language :", font="Times 13 bold", background="White")
splang.grid(row=4, column=2, sticky=W, pady=10)
cbomt3 = ttk.Combobox(tf2, font="Arial 12 bold", state='readonly', width=23)
cbomt3['value'] = ('', 'English', 'Hindi.', 'Marathi', 'Tamil', 'Gujrati', 'Kannada', 'Bangali', 'Panjabi')
cbomt3.current(0)
cbomt3.grid(row=5, column=2, sticky=W, pady=20)

birth_d = Label(tf2, text="Enter Your Birth Date :", font="Times 13 bold", background="White")
birth_d.grid(row=6, column=1, sticky=W, padx=20, pady=10)
bde3 = Entry(tf2, font="Arial 12 bold", textvariable=birthdate, width=23)
bde3.grid(row=7, column=1, pady=20)

age_l = Label(tf2, text="Enter Your Age :", font="Times 13 bold", background="White")
age_l.grid(row=6, column=0, sticky=W, pady=10)
cbomage = ttk.Combobox(tf2, font="Arial 12 bold", state='readonly', textvariable=age, width=23)
cbomage['value'] = (
    '', '1 - 10', '11 - 20', '21 - 30', '31 - 40', '41 - 50', '51 - 60', '61 - 70', '71 - 80', '81 - 90', '91 - 100')
cbomage.current(0)
cbomage.grid(row=7, column=0, pady=10)

lbldate = Label(tf2, text="Date", font='arial 12 bold', background="White")
lbldate.grid(row=6, column=2, sticky=W, )
Entdate = Entry(tf2, font='arial 12 bold', textvariable=Date, width=24)
Entdate.grid(row=7, column=2, sticky=W, pady=20)

# Address Details :

addr_lbl = Label(tf2, text=" Address Details : ", font="Times 16 bold", background="White")
addr_lbl.grid(row=1, column=3, sticky=W, padx=25, pady=20)

country_l = Label(tf2, text="Enter country :", font="Times 13 bold", background="White")
country_l.grid(row=2, column=3, sticky=W, padx=25, pady=35)
country_e = Entry(tf2, font="Arial 12 bold", width=23)
country_e.grid(row=3, column=3, padx=25, pady=30)

state_l = Label(tf2, text="Enter State :", font="Times 13 bold", background="White")
state_l.grid(row=4, column=3, sticky=W, padx=25, pady=25)
state_e = Entry(tf2, font="Arial 12 bold", width=23)
state_e.grid(row=5, column=3, padx=25, pady=30)

city_l = Label(tf2, text="Enter city:", font="Helvetica 13 bold", background="White")
city_l.grid(row=6, column=3, sticky=W, padx=25, pady=25)
city_e = Entry(tf2, font="Arial 12 bold", width=23)
city_e.grid(row=7, column=3, padx=25, pady=30)

addrel = Label(tf2, text="Enter Address:", font="Times 13 bold", background="White")
addrel.grid(row=2, column=4, sticky=W, padx=25, pady=25)
addree = Entry(tf2, font="Arial 12 bold", textvariable=address, width=23)
addree.grid(row=3, column=4, padx=25, pady=30)

pin_codel = Label(tf2, text="Enter Pin Code :", font="Times 13 bold", background="White")
pin_codel.grid(row=4, column=4, sticky=W, padx=25, pady=25)
pin_codee = Entry(tf2, font="Arial 12 bold", textvariable=pin_code, width=23)
pin_codee.grid(row=5, column=4, padx=25, pady=30)

# contact details :

contact_lbl = Label(tf2, text=" Contact Details : ", font="Times  16 bold", background="White")
contact_lbl.grid(row=1, column=5, sticky=W, padx=5, pady=15)

co_no = Label(tf2, text="Enter contact number :", font="Times 13 bold", background="White")
co_no.grid(row=2, column=5, sticky=W, pady=25)
co_noe = Entry(tf2, font="Arial 12 bold", textvariable=contact, width=23)
co_noe.grid(row=3, column=5, pady=30)

mail = Label(tf2, text="Enter Email-Id :", font="Times 13 bold", background="White")
mail.grid(row=4, column=5, sticky=W, pady=25)
mail_e = Entry(tf2, font="Arial 12 bold", textvariable=email, width=23)
mail_e.grid(row=5, column=5, pady=30)

submit_button = Button(tf2, text="Submit ", font="Times 14 bold", background="navy", foreground="White",
                       command=next_page, width=18)
submit_button.grid(row=7, column=4, padx=10, pady=25)

exit_btn = Button(tf2, text=" EXIT ", font="Times 14 bold", background="navy", foreground="White", command=exit2,
                  width=18)
exit_btn.grid(row=7, column=5, padx=10, pady=30)

root.mainloop()

