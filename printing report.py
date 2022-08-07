import openpyxl
import time
import os
from win32com import client

wb = openpyxl.load_workbook("D:\\Documents\\Python Projects\\CODVID-19 Prediction Project\\Report_Sample_1.xlsx")
sh1 = wb['sheet1']
row = sh1.max_row
column = sh1.max_column




Date = 'anything'


sh1.cell(row=7, column=3, value=Date)

wb.save("C:\\Users\\Swarup\\Desktop\\sampleoutput_1.xlsx", )
xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open('C:\\Users\\Swarup\\Desktop\\sampleoutput_1.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.ExportAsFixedFormat(0, 'C:\\Users\\Swarup\\Desktop\\sampleoutput_1.pdf')
